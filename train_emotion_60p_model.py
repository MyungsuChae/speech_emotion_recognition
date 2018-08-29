
import numpy as np
import os
import pandas as pd
import keras
import keras.layers
from keras.models import load_model, Model
from keras.callbacks import EarlyStopping
from keras.models import Sequential
from keras.layers import Dense, Dropout
from keras.layers import Embedding
from keras.layers import Conv1D, GlobalAveragePooling1D, MaxPooling1D, GlobalMaxPooling1D, AveragePooling1D, AveragePooling2D, MaxPooling2D, concatenate, Concatenate, Input, Flatten, Activation
from keras.optimizers import adam
from keras.utils.training_utils import multi_gpu_model
from keras.layers.normalization import BatchNormalization
from keras import regularizers
import tensorflow as tf
from sklearn.model_selection import train_test_split
import time
import openpyxl

def init():
    ### GPU Setting
    os.environ["CUDA_VISIBLE_DEVICES"]="1"

### Load Data

def load_data():
    file_path = "/data3/mschae89/FlagshipAER/"
    wb = openpyxl.load_workbook("/data3/younghoon/Audio_data/voice_annotation_emotion.xlsx")
    ws = wb.get_sheet_by_name("Sheet1")

    Num_data = 9500
    Num_seq  = 1000
    Num_feat = 80
    Num_class = 7
    np.random.seed(5)
    x_data = np.zeros((Num_data, Num_seq, Num_feat))
    y_data = np.zeros(Num_data)

    for i in range(0, Num_data):
        audio_name = ws.cell(row=i+2, column=1).value
        df = pd.read_csv(file_path + "mels_40_squared/" + audio_name + "_mels.csv")
        x_data[i, :, :] = df
        y_data[i] = ws.cell(row=i+2, column=6).value  # column = 2: Acryl Emotion (6 class), 6: KI4AI Emotion (7 class), KI4AI Gender Emotion (14 class)

    ### Data seperation
    x_train, x_test, y_train, y_test = train_test_split(x_data[:Num_data,:,:], y_data, test_size = 0.1, random_state = 42, shuffle=False)
    x_train, x_val, y_train, y_val = train_test_split(x_train, y_train, test_size = 0.1, random_state = 42, shuffle=False)


    ### Convert labels to categorical one-hot encoding
    y_train = keras.utils.to_categorical(y_train, num_classes=Num_class)
    y_val  = keras.utils.to_categorical(y_val, num_classes=Num_class)
    y_test = keras.utils.to_categorical(y_test, num_classes=Num_class)
    print('Data Load Done')

    return x_train, x_val, x_test, y_train, y_val, y_test, Num_class


### Model build

def ConvMax(model_input, filter_size, kernel_size):

    model = Conv1D(filter_size, kernel_size)(model_input)
    #model = BatchNormalization()(model)
    model = Activation('relu')(model)
    model = MaxPooling1D(pool_size=2, strides=2, padding='valid')(model)

    return model


def ConvAverage(model_input, filter_size, kernel_size):

    model = Conv1D(filter_size, kernel_size)(model_input)
    #model = BatchNormalization()(model)
    model = Activation('relu')(model)
    model = AveragePooling1D(pool_size=2, strides=2, padding='valid')(model)

    return model


def emotion_model(seq_length, input_length, nb_classes):

    dropout = 0.25

    model_in = Input(shape=(seq_length, input_length))

    model = ConvMax(model_in, 512, 5)

    model1 = ConvMax(model, 512, 5)
    model1 = ConvMax(model1, 512, 5)
    # model1 = ConvMax(model1, 512, 5)
    model1 = ConvMax(model1, 512, 5)
    model1 = ConvMax(model1, 512, 5)
    model1 = GlobalMaxPooling1D()(model1)
    model1 = Dropout(dropout)(model1)


    model2 = ConvAverage(model, 512, 5)
    model2 = ConvAverage(model2, 512, 5)
    model2 = ConvAverage(model2, 512, 5)
    model2 = ConvAverage(model2, 512, 5)
    model2 = ConvAverage(model2, 512, 5)
    model2 = GlobalAveragePooling1D()(model2)
    model2 = Dropout(dropout)(model2)

    model_c = concatenate([model1, model2])

    model_c = Dense(512, activation='relu', kernel_regularizer=regularizers.l2(0.01))(model_c)
    model_c = Dropout(dropout)(model_c)

    model_out = Dense(nb_classes, activation='softmax', name='output_layer')(model_c)
    model = Model(inputs=model_in, outputs=model_out)

#    model_con = Dense(300, activation='relu', name='dense_2')(model_con)
#    model = Dropout(dropout)(model)
    print(model.summary())
    ### Train model
    return model

def main():
    init()
    x_train, x_val, x_test, y_train, y_val, y_test, Num_class = load_data()

    print('Data Load Completed!')
    ### Model build
    model = emotion_model( seq_length=np.size(x_test,1), input_length=np.size(x_test,2), nb_classes= Num_class ) ### NEED TO CHANGE
    ### Pramameter setting
    gpu_options = tf.GPUOptions(per_process_gpu_memory_fraction=0.9)
    config = tf.ConfigProto(device_count={'GPU': 1, 'CPU': 30},gpu_options=gpu_options)
    sess = tf.Session(config=config)
    keras.backend.set_session(sess)

    # adam1=keras.optimizers.Adam(lr = 0.01, beta_1 = 0.9, beta_2 = 0.999, epsilon=None, decay=0.9, amsgrad=False)
    # model.compile(loss='categorical_crossentropy', optimizer=adam1, metrics=['accuracy'])
    # model.fit(x_train, y_train, batch_size=500, epochs=100, validation_data= (x_val, y_val), verbose=2)

    # early_stopping = EarlyStopping(monitor='val_acc', min_delta = 0.01, patience = 100, verbose = 1, mode='auto')

    adam2=keras.optimizers.Adam(lr = 0.001, beta_1 = 0.9, beta_2 = 0.999, epsilon=None, decay=0, amsgrad=False)
    model.compile(loss='categorical_crossentropy', optimizer=adam2, metrics=['accuracy'])
    model.fit(x_train, y_train, batch_size=500, epochs=200, validation_data= (x_val, y_val), verbose=2)
    ### Score the model
    score = model.evaluate(x_test, y_test, batch_size=10)
    print('Test score:', score[0])
    print('Test accuracy:', score[1])


    ### Save model
    file_path = "/data3/mschae89/FlagshipAER/"
    model_name = 'Emotion_model_mels40_'+ str(time.localtime().tm_mon).zfill(2) + str(time.localtime().tm_mday).zfill(2) + '_' + str(time.localtime().tm_hour).zfill(2) + str(time.localtime().tm_min).zfill(2)
    model_name = file_path + 'model/' + model_name + '_MA_' + str(round(score[1]*100)) +'.h5'
    print(model_name)
    model.save(model_name)


if __name__ == '__main__':
    main()