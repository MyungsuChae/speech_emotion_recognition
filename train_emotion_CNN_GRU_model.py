#from __future__ import absolute_import
#from __future__ import division
#from __future__ import print_function

import numpy as np
import os
import pandas as pd
import keras
import keras.layers
from keras.models import load_model, Model
from keras.callbacks import EarlyStopping
from keras.models import Sequential
from keras.layers import Dense, Dropout

from keras.layers import Conv1D, GlobalAveragePooling1D, MaxPooling1D, GlobalMaxPooling1D, AveragePooling1D, AveragePooling2D, MaxPooling2D, concatenate, Concatenate, Input, Flatten, Activation, GRU, TimeDistributed, LSTM, Add
from keras.optimizers import adam
from keras.utils.training_utils import multi_gpu_model
from keras.layers.normalization import BatchNormalization
from keras import regularizers
import tensorflow as tf
from sklearn.model_selection import train_test_split
import time
import openpyxl

def init():
    ### GPU Setting
    os.environ["CUDA_VISIBLE_DEVICES"]="2"

### Load Data
def load_data():
    file_path = "NEED_TO_CHANGE"
    wb = openpyxl.load_workbook("NEED_TO_CHANGE")
    ws = wb.get_sheet_by_name("NEED_TO_CHANGE")

    Num_data_a = 9500
    Num_data_k = 0
#    Num_data_k = 2714
    Num_seq  = 1000
    Num_feat = 40
    Num_class = 7
    np.random.seed(5)
    x_data = np.zeros((Num_data_a+Num_data_k, Num_seq, Num_feat))
    y_data = np.zeros(Num_data_a+Num_data_k)
    x_val = np.zeros((Num_data_k, Num_seq, Num_feat))
    y_val = np.zeros(Num_data_k)
    for i in range(0, Num_data_a):
        audio_name = ws.cell(row=i+2, column=1).value
        df = pd.read_csv(file_path + "mels_40a/" + audio_name + "_mels.csv")
        x_data[i, :, :] = df
        y_data[i] = ws.cell(row=i+2, column=6).value  # 6: KI4AI Emotion (7 class), KI4AI Gender Emotion (14 class)



    ### Data seperation
    x_train, x_test, y_train, y_test = train_test_split(x_data, y_data, test_size = 0.1, random_state = 42, shuffle=False)
    x_train, x_val, y_train, y_val = train_test_split(x_train, y_train, test_size = 0.1, random_state = 42, shuffle=False)

    ### Convert labels to categorical one-hot encoding
    y_train = keras.utils.to_categorical(y_train, num_classes=Num_class)
    y_test = keras.utils.to_categorical(y_test, num_classes=Num_class)
    #y_val  = keras.utils.to_categorical(y_val, num_classes=Num_class)
    print('Data Load Done')

    return x_train, x_val, x_test, y_train, y_val, y_test, Num_class


def ConvMax(model_input, filter_size, kernel_size):
    model = Conv1D(filter_size, kernel_size, activation='relu', padding='valid')(model_input)
    model = MaxPooling1D(pool_size=2, strides=2, padding='valid')(model)

    return model

def ConvAverage(model_input, filter_size, kernel_size):
    model = Conv1D(filter_size, kernel_size, activation='relu', padding='valid')(model_input)
    model = AveragePooling1D(pool_size=2, strides=2, padding='valid')(model)
    return model

def emotion_model(seq_length, input_length, nb_classes):

    dropout = 0.25

    model_in = Input(shape=(seq_length, input_length))

    model = ConvMax(model_in, 512, 5)
    model = Dropout(dropout)(model)
    model1 = ConvMax(model, 512, 5)
    model1 = Dropout(dropout)(model1)
    model1 = ConvMax(model1, 512, 5)
    model1 = Dropout(dropout)(model1)

    model2 = ConvAverage(model, 512, 5)
    model2 = Dropout(dropout)(model2)
    model2 = ConvAverage(model2, 512, 5)
    model2 = Dropout(dropout)(model2)

    model3_1 = GRU(512, return_sequences=True, dropout = dropout)(model1)
    model3_1 = GlobalMaxPooling1D()(model3_1)

    model3_2 = GRU(512, return_sequences=True, dropout = dropout)(model2)
    model3_2 = GlobalMaxPooling1D()(model3_2)

    model4 = concatenate([model3_1, model3_2])

    model_out = Dense(512, activation='relu', kernel_regularizer=regularizers.l2(0.01))(model4)
    model_out = Dropout(dropout)(model_out)

    model_out = Dense(512, activation='relu', kernel_regularizer=regularizers.l2(0.01))(model_out)
    model_out = Dropout(dropout)(model_out)
    model_out = Dense(nb_classes, activation='softmax', name='output_layer')(model_out)

    model = Model(inputs=model_in, outputs=model_out)

    print(model.summary())

    ### Train model
    return model

def main():
    init()
    x_train, x_val, x_test, y_train, y_val, y_test, Num_class = load_data()

    print('Data Load Completed!')

    ### Model build
    model = emotion_model( seq_length=np.size(x_test,1), input_length=np.size(x_test,2), nb_classes= Num_class ) ### NEED TO CHANGE

    ### Pramameter setting
    gpu_options = tf.GPUOptions(per_process_gpu_memory_fraction=0.9)
    config = tf.ConfigProto(device_count={'GPU': 1, 'CPU': 30},gpu_options=gpu_options)
    sess = tf.Session(config=config)
    keras.backend.set_session(sess)

    ### Train

    early_stopping = EarlyStopping(monitor='val_acc', min_delta = 0.001, patience = 20, verbose = 1, mode='auto')

    adam2=keras.optimizers.Adam(lr = 0.001, beta_1 = 0.9, beta_2 = 0.999, epsilon=None, decay=0, amsgrad=False)
    model.compile(loss='categorical_crossentropy', optimizer=adam2, metrics=['accuracy'])
    model.fit(x_train, y_train, batch_size=300, epochs=200, validation_data= (x_val, y_val), verbose=2, callbacks=[early_stopping])

    ### Evaluation
    score = model.evaluate(x_test, y_test, batch_size=100)
    print('Test score:', score[0])
    print('Test accuracy:', score[1])

    ### Save
    file_path = "NEED_TO_CHANGE"
    model_name = file_path + 'model/' + 'Emotion_model_'+ str(time.localtime().tm_mon).zfill(2) + str(time.localtime().tm_mday).zfill(2) + '_' + str(time.localtime().tm_hour).zfill(2) + str(time.localtime().tm_min).zfill(2)
    model_name = model_name + '_' + str(round(score[1]*100)) +'.h5'
    print(model_name)
    model.save(model_name)

if __name__ == '__main__':
    main()
